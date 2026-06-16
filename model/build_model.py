#!/usr/bin/env python3
"""
HCTECH — Modèle financier 7 ans — MACHINES EN LEASING.
Chaque machine est financée à 100 % par leasing (~12 %/7 ans) et s'autofinance
par ses propres revenus. Capital social (300 k, Nessim) = fonds de roulement.
Génère : simulation mensuelle -> agrégats annuels -> classeur Excel investisseur.
"""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# ════════════════════════ HYPOTHÈSES ════════════════════════
MONTHS = 84
USD_TND = 2.95
MACHINE_USD = 18000
MACHINE_DT = MACHINE_USD * USD_TND
IMPORT_FEES = MACHINE_DT * 0.025
INSTALL = 3500
MACHINE_HT = MACHINE_DT + IMPORT_FEES + INSTALL          # 57 928

# Leasing machines
LEASE_RATE = 0.12
LEASE_YEARS = 7
LEASE_N = LEASE_YEARS * 12
MACHINE_LEASE_MONTH = MACHINE_HT * (LEASE_RATE/12) / (1 - (1 + LEASE_RATE/12)**(-LEASE_N))
MACHINE_LEASE_YEAR = MACHINE_LEASE_MONTH * 12

# Revenu
STATION_L_DAY = 12000
RECOVERY = 5/1000
L_MACHINE_DAY = STATION_L_DAY * RECOVERY                 # 60 L/j
L_MACHINE_YEAR = L_MACHINE_DAY * 365                     # 21 900 L/an
FUEL_P0 = 2.525
FUEL_GROWTH = 0.045
HCTECH_SHARE = 0.60
STATION_SHARE = 0.40

# Charges
OPEX_MACHINE_MONTH = 120
EMP_CHARGES = 0.21
SAL_FOUNDER0 = 2000
SAL_FOUNDER_GROWTH = 0.50
SAL_FOUNDER_CAP = 6000
SAL_TECH0 = 1200
SAL_TECH_GROWTH = 0.07
MACHINES_PER_TECH = 48
RENT0 = 1200
RENT_GROWTH = 0.06
OVERHEAD0 = 2000
OVERHEAD_GROWTH = 0.06

# Leasing véhicule
VEH_PRICE = 70000
VEH_RATE = 0.12
VEH_MONTHS = 60
VEH_PMT = VEH_PRICE * (VEH_RATE/12) / (1 - (1 + VEH_RATE/12)**(-VEH_MONTHS))

# Fiscalité
IS_RATE = 0.25
CSS_RATE = 0.03
TAX_RATE = IS_RATE + CSS_RATE
MIN_TAX = 0.001

# Financement
CAPITAL = 300000
DIVIDEND_PAYOUT = 0.0      # bonus/dividendes actionnaires (à déterminer) — 0 par défaut

# Déploiement: 14 machines à M3 puis +14 tous les 6 mois
TARGET_BATCH = 14
DEPLOY_MONTHS = list(range(3, MONTHS + 1, 6))

# Cap table (sortie de Nazeh — ses parts cédées à Nessim et Lamine, juin 2026)
SHARES = {'Nessim Mami': 0.50, 'Ali Ben Hamoud': 0.25, 'Lamine': 0.25}
CASH_CONTRIB = {'Nessim Mami': 300000, 'Ali Ben Hamoud': 0, 'Lamine': 0}

# ════════════════════════ HELPERS ════════════════════════
def year_of(m): return (m - 1)//12 + 1
def fuel_price_y(y): return FUEL_P0 * (1+FUEL_GROWTH)**(y-1)
def founder_salary_y(y): return min(SAL_FOUNDER0*(1+SAL_FOUNDER_GROWTH)**(y-1), SAL_FOUNDER_CAP)
def tech_unit_y(y): return SAL_TECH0*(1+SAL_TECH_GROWTH)**(y-1)
def rent_y(y): return RENT0*(1+RENT_GROWTH)**(y-1)
def overhead_y(y): return OVERHEAD0*(1+OVERHEAD_GROWTH)**(y-1)

def machine_rev_month(y):
    return L_MACHINE_DAY*(365/12)*fuel_price_y(y)*HCTECH_SHARE

# ════════════════════════ SIMULATION MENSUELLE ════════════════════════
def simulate():
    deploy = {dm: TARGET_BATCH for dm in DEPLOY_MONTHS}     # leasing -> plan = cible
    def mis_at(mo): return sum(n for dm,n in deploy.items() if dm <= mo)

    cash = 0.0
    tax_due_next = 0.0
    min_cash = float('inf')
    yacc = {y: dict(rev=0,opex=0,payroll=0,rent=0,overhead=0,veh=0,mlease=0,
                    new=0, techs=0) for y in range(1,8)}
    cash_year_end = {}
    monthly_rows = []

    for m in range(1, MONTHS+1):
        y = year_of(m)
        if m == 1:
            cash += CAPITAL
        mis = mis_at(m)
        techs = max(1, math.ceil(mis/MACHINES_PER_TECH))
        rev = mis * machine_rev_month(y)
        opex = mis * OPEX_MACHINE_MONTH
        gross_pay = 2*founder_salary_y(y) + techs*tech_unit_y(y)
        payroll = gross_pay*(1+EMP_CHARGES)
        veh = VEH_PMT if m <= VEH_MONTHS else 0.0
        mlease = mis * MACHINE_LEASE_MONTH
        rentm = rent_y(y); ovh = overhead_y(y)

        tax_cash = 0.0
        if m % 12 == 6:
            tax_cash = tax_due_next; tax_due_next = 0.0

        cash += rev - opex - payroll - veh - mlease - rentm - ovh - tax_cash
        min_cash = min(min_cash, cash)

        yacc[y]['rev'] += rev; yacc[y]['opex'] += opex; yacc[y]['payroll'] += payroll
        yacc[y]['rent'] += rentm; yacc[y]['overhead'] += ovh; yacc[y]['veh'] += veh
        yacc[y]['mlease'] += mlease; yacc[y]['techs'] = techs

        monthly_rows.append(dict(m=m,y=y,mis=mis,rev=rev,cash=cash))

        if m % 12 == 0:
            a = yacc[y]
            ebitda = a['rev']-a['opex']-a['payroll']-a['rent']-a['overhead']-a['veh']-a['mlease']
            tax = ebitda*TAX_RATE if ebitda>0 else MIN_TAX*a['rev']
            tax_due_next = tax
            a['ebitda']=ebitda; a['tax']=tax; a['net']=ebitda-tax
            a['new']= sum(n for dm,n in deploy.items() if year_of(dm)==y)
            a['parc']= mis_at(m)
            cash_year_end[y]=cash

    return yacc, cash_year_end, min_cash, monthly_rows, deploy

YA, CASH_EOY, MIN_CASH, MROWS, DEPLOY = simulate()

# machine-years par an (pour transparence)
def machine_years(y):
    # somme des mois en service / 12 sur l'année
    tot=0
    for m in range((y-1)*12+1, y*12+1):
        tot += sum(n for dm,n in DEPLOY.items() if dm<=m)
    return tot/12
MY = {y: machine_years(y) for y in range(1,8)}

# ── Dividendes (Option 1 : Nessim prioritaire capital+20% = 360k, puis pro-rata, payout 70%) ──
PREF_TARGET = 360000
PAYOUT = 0.70
def compute_dividends():
    flows={y:{n:0 for n in SHARES} for y in range(1,8)}; rec=0
    for y in range(1,8):
        pool=PAYOUT*YA[y]['net']
        if rec<PREF_TARGET:
            pay=min(pool,PREF_TARGET-rec); flows[y]['Nessim Mami']+=pay; rec+=pay; pool-=pay
        for n,p in SHARES.items(): flows[y][n]+=pool*p
    return flows
DIV=compute_dividends()
DIV_TOTAL={y:sum(DIV[y].values()) for y in range(1,8)}
SH_TOTAL={n:sum(DIV[y][n] for y in range(1,8)) for n in SHARES}
# trésorerie après distribution (dividende déclaré en y, versé en y+1)
CASH_AFTER={}; cum_paid=0
for y in range(1,8):
    CASH_AFTER[y]=CASH_EOY[y]-cum_paid
    cum_paid+=DIV_TOTAL[y]
def _irr(cf):
    lo,hi=-0.99,10.0
    for _ in range(200):
        mid=(lo+hi)/2
        npv=sum(c/((1+mid)**t) for t,c in enumerate(cf))
        lo,hi=(mid,hi) if npv>0 else (lo,mid)
    return (lo+hi)/2
NESSIM_IRR=_irr([-CAPITAL]+[DIV[y]['Nessim Mami'] for y in range(1,8)])
cum=0; NESSIM_PAYBACK=7
for y in range(1,8):
    cum+=DIV[y]['Nessim Mami']
    if cum>=CAPITAL: NESSIM_PAYBACK=y; break

# ════════════════════════ EXPORT EXCEL ════════════════════════
DARK="0D1F17"; GREEN="1A7A4A"; GREEN2="2ECC71"; LIGHT="E8F7EF"; GREY="F4F7F5"; AMBER="F39C12"; RED="E74C3C"
thin = Side(style="thin", color="D0E4D8")
border = Border(left=thin,right=thin,top=thin,bottom=thin)

def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
def title(ws, text, sub=""):
    ws.merge_cells("A1:I1")
    c=ws["A1"]; c.value=text; c.font=Font(bold=True,size=16,color=GREEN)
    if sub:
        ws.merge_cells("A2:I2")
        ws["A2"].value=sub; ws["A2"].font=Font(italic=True,size=10,color="5A7060")
DTFMT='#,##0 "DT"'; DTFMT2='#,##0'; PCT='0.0%'

wb = Workbook()

# ---- Sheet 1: Synthèse ----
ws = wb.active; ws.title="Synthèse"
title(ws,"HCTECH — Business Plan 7 ans","Modèle Revenue Share 60/40 · Machines en leasing · Autofinancement par les revenus")
ws["A4"]="INDICATEURS CLÉS"; ws["A4"].font=Font(bold=True,size=12,color=DARK)
kpis=[
 ("Parc de machines à 7 ans", f"{YA[7]['parc']:.0f} machines"),
 ("CA HCTECH année 7", f"{YA[7]['rev']:,.0f} DT"),
 ("Résultat net année 7", f"{YA[7]['net']:,.0f} DT"),
 ("Résultat net cumulé 7 ans", f"{sum(YA[y]['net'] for y in range(1,8)):,.0f} DT"),
 ("Trésorerie fin année 7", f"{CASH_EOY[7]:,.0f} DT"),
 ("Trésorerie minimale (84 mois)", f"{MIN_CASH:,.0f} DT"),
 ("Capital social (Nessim)", f"{CAPITAL:,.0f} DT"),
 ("Marge nette / machine / an (rég.)", f"{machine_rev_month(1)*12 - MACHINE_LEASE_YEAR - OPEX_MACHINE_MONTH*12:,.0f} DT"),
 ("Gain station / mois (an 1)", f"{L_MACHINE_YEAR*FUEL_P0*STATION_SHARE/12:,.0f} DT"),
 ("Coût machine tout compris (HT)", f"{MACHINE_HT:,.0f} DT"),
 ("Leasing machine / an", f"{MACHINE_LEASE_YEAR:,.0f} DT"),
 ("— Dividendes (Option 1) —", "Nessim prioritaire"),
 ("Capital Nessim remboursé", f"fin An {NESSIM_PAYBACK}"),
 ("TRI Nessim", f"{NESSIM_IRR*100:,.0f} %"),
 ("Nessim — total reçu 7 ans", f"{SH_TOTAL['Nessim Mami']:,.0f} DT"),
 ("Dividendes distribués 7 ans", f"{sum(DIV_TOTAL.values()):,.0f} DT"),
 ("Trésorerie après distrib. An 7", f"{CASH_AFTER[7]:,.0f} DT"),
]
r=5
for k,v in kpis:
    ws[f"A{r}"]=k; ws[f"A{r}"].font=Font(size=11)
    ws[f"C{r}"]=v; ws[f"C{r}"].font=Font(bold=True,size=11,color=GREEN)
    r+=1
ws.column_dimensions["A"].width=34; ws.column_dimensions["C"].width=24

# ---- Sheet 2: Hypothèses ----
ws=wb.create_sheet("Hypothèses")
title(ws,"Hypothèses du modèle","Toutes les hypothèses externes sont sourcées (juin 2026) — voir BP")
heads=["Paramètre","Valeur","Unité","Source / justification"]
ws.append([]); ws.append([]); ws.append(heads)
for i,h in enumerate(heads,1): style_header(ws.cell(row=3,column=i))
rows=[
 ("Taux de change USD/DT",USD_TND,"DT/USD","BCT ~2,94 (juin 2026), retenu 2,95 prudent"),
 ("Prix machine départ usine",MACHINE_USD,"USD","Fournisseur GEKO Corée"),
 ("Prix machine en DT",MACHINE_DT,"DT","18 000 × 2,95"),
 ("Frais d'import",IMPORT_FEES,"DT","2,5 % du prix machine"),
 ("Installation",INSTALL,"DT","Prestataire local"),
 ("Machine tout compris (HT)",MACHINE_HT,"DT","Immobilisation / base leasing. TVA 6 % récupérable"),
 ("Taux leasing machine",LEASE_RATE,"%/an","Leasing Tunisie ~12 % (plafond usure 16,28 %)"),
 ("Durée leasing machine",LEASE_YEARS,"ans","Aligné durée de vie / contrat"),
 ("Leasing machine / mois",MACHINE_LEASE_MONTH,"DT","Mensualité constante"),
 ("Ventes station moyenne",STATION_L_DAY,"L/jour","Station cible ≥ 12 000 L/j"),
 ("Taux de récupération",RECOVERY,"L/L","5/1000 — validé prototype 2019"),
 ("Essence récupérée / machine",L_MACHINE_YEAR,"L/an","60 L/j × 365"),
 ("Prix essence an 1",FUEL_P0,"DT/L","SNDP mai 2025"),
 ("Hausse prix essence",FUEL_GROWTH,"%/an","CAGR réel 2016-2026 ≈ +4,6 %"),
 ("Part HCTECH",HCTECH_SHARE,"%","Partage 60/40"),
 ("Part station",STATION_SHARE,"%","Partage 60/40"),
 ("OPEX machine (maint+internet)",OPEX_MACHINE_MONTH,"DT/mois","Par machine"),
 ("Charges patronales",EMP_CHARGES,"%","CNSS 17,07 + TFP 2 + FOPROLOS 1 + acc."),
 ("Salaire fondateur an 1",SAL_FOUNDER0,"DT/mois","Lamine & Ali"),
 ("Croissance salaire fondateur",SAL_FOUNDER_GROWTH,"%/an","Plafonné à 6 000 DT/mois"),
 ("Plafond salaire fondateur",SAL_FOUNDER_CAP,"DT/mois","Décision actionnaires"),
 ("Salaire technicien an 1",SAL_TECH0,"DT/mois","+7 %/an, 1 pour 48 machines"),
 ("Loyer bureau+atelier an 1",RENT0,"DT/mois","+6 %/an"),
 ("Frais généraux an 1",OVERHEAD0,"DT/mois","Compta, assur., banque, carburant… (+6 %/an)"),
 ("Leasing véhicule / mois",VEH_PMT,"DT","70 000 DT @ 12 %/5 ans"),
 ("IS (impôt sociétés)",IS_RATE,"%","Services/commerce — LF 2025"),
 ("CSS (contrib. solidarité)",CSS_RATE,"%","LF 2026, sociétés < 35 %"),
 ("Minimum d'impôt",MIN_TAX,"% du CA","Si déficit"),
 ("Capital social",CAPITAL,"DT","Apporté intégralement par Nessim"),
]
for row in rows:
    ws.append(list(row))
for rr in range(4,4+len(rows)):
    ws.cell(row=rr,column=1).font=Font(size=10)
    vcell=ws.cell(row=rr,column=2)
    unit=ws.cell(row=rr,column=3).value
    if unit=="%" or unit=="%/an" or unit=="L/L" or unit=="% du CA":
        vcell.number_format=PCT
    else:
        vcell.number_format=DTFMT2
    vcell.font=Font(bold=True,color=GREEN)
    for cc in range(1,5): ws.cell(row=rr,column=cc).border=border
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=14
ws.column_dimensions["C"].width=10; ws.column_dimensions["D"].width=52

# ---- Sheet 3: Déploiement ----
ws=wb.create_sheet("Déploiement")
title(ws,"Déploiement du parc","14 machines à M3, puis +14 tous les 6 mois — financées en leasing")
ws.append([]); ws.append([])
hdr=["(année)"]+[f"An {y}" for y in range(1,8)]
ws.append(hdr)
for i in range(1,9): style_header(ws.cell(row=3,column=i))
def add_row(label, vals, fmt=DTFMT2, bold=False, fill=None):
    ws.append([label]+list(vals))
    r=ws.max_row
    lc=ws.cell(row=r,column=1); lc.font=Font(bold=bold,size=10)
    for j in range(2,9):
        c=ws.cell(row=r,column=j); c.number_format=fmt; c.border=border
        if bold: c.font=Font(bold=True)
        if fill: c.fill=PatternFill("solid",fgColor=fill)
    lc.border=border
    if fill: lc.fill=PatternFill("solid",fgColor=fill)
add_row("Nouvelles machines",[YA[y]['new'] for y in range(1,8)])
add_row("Parc fin d'année",[YA[y]['parc'] for y in range(1,8)],bold=True,fill=LIGHT)
add_row("Machine-années (prod.)",[round(MY[y],1) for y in range(1,8)],fmt='0.0')
add_row("Techniciens",[YA[y]['techs'] for y in range(1,8)],fmt='0')
add_row("Prix essence (DT/L)",[round(fuel_price_y(y),3) for y in range(1,8)],fmt='0.000')
for col in range(1,9): ws.column_dimensions[get_column_letter(col)].width=15

# ---- Sheet 4: Compte de résultat ----
ws=wb.create_sheet("Compte de résultat")
title(ws,"Compte de résultat prévisionnel (7 ans)","En DT — machines en leasing (pas d'amortissement machine, le leasing porte le financement)")
ws.append([]); ws.append([])
ws.append(["(en DT)"]+[f"An {y}" for y in range(1,8)])
for i in range(1,9): style_header(ws.cell(row=3,column=i))
add_row("CA HCTECH (60 %)",[YA[y]['rev'] for y in range(1,8)],bold=True)
add_row("(−) OPEX machines",[-YA[y]['opex'] for y in range(1,8)])
add_row("(−) Leasing machines",[-YA[y]['mlease'] for y in range(1,8)])
add_row("(−) Masse salariale",[-YA[y]['payroll'] for y in range(1,8)])
add_row("(−) Loyer",[-YA[y]['rent'] for y in range(1,8)])
add_row("(−) Frais généraux",[-YA[y]['overhead'] for y in range(1,8)])
add_row("(−) Leasing véhicule",[-YA[y]['veh'] for y in range(1,8)])
add_row("Résultat d'exploitation",[YA[y]['ebitda'] for y in range(1,8)],bold=True,fill=LIGHT)
add_row("(−) Impôt (IS 25 % + CSS 3 %)",[-YA[y]['tax'] for y in range(1,8)])
add_row("RÉSULTAT NET",[YA[y]['net'] for y in range(1,8)],bold=True,fill=LIGHT)
cum=0; cums=[]
for y in range(1,8):
    cum+=YA[y]['net']; cums.append(cum)
add_row("Résultat net cumulé",cums,bold=True)
for col in range(1,9): ws.column_dimensions[get_column_letter(col)].width=16
ws.column_dimensions["A"].width=28

# ---- Sheet 5: Trésorerie ----
ws=wb.create_sheet("Trésorerie")
title(ws,"Trésorerie (fin d'année)","Capital 300 k injecté an 1 — le parc s'autofinance, trésorerie toujours positive")
ws.append([]); ws.append([])
ws.append(["(en DT)"]+[f"An {y}" for y in range(1,8)])
for i in range(1,9): style_header(ws.cell(row=3,column=i))
add_row("Résultat net",[YA[y]['net'] for y in range(1,8)])
add_row("Trésorerie avant distribution",[CASH_EOY[y] for y in range(1,8)])
add_row("(−) Dividendes versés (cumulés)",[CASH_EOY[y]-CASH_AFTER[y] for y in range(1,8)])
add_row("Trésorerie après distribution",[CASH_AFTER[y] for y in range(1,8)],bold=True,fill=LIGHT)
ws.append([])
ws.append([f"Trésorerie minimale sur 84 mois (avant distrib.) : {MIN_CASH:,.0f} DT"])
ws.cell(row=ws.max_row,column=1).font=Font(bold=True,color=GREEN if MIN_CASH>0 else RED)
for col in range(1,9): ws.column_dimensions[get_column_letter(col)].width=16
ws.column_dimensions["A"].width=28

# ---- Sheet 6: Économie par machine ----
ws=wb.create_sheet("Unité par machine")
title(ws,"Économie unitaire par machine (an 1)","Chaque machine est bénéficiaire dès le 1er mois")
ws.append([]); ws.append([])
unit=[
 ("Essence récupérée",f"{L_MACHINE_YEAR:,.0f} L/an","60 L/j × 365"),
 ("Valeur brute récupérée",f"{L_MACHINE_YEAR*FUEL_P0:,.0f} DT/an","× 2,525 DT"),
 ("Part HCTECH (60 %)",f"{L_MACHINE_YEAR*FUEL_P0*HCTECH_SHARE:,.0f} DT/an",""),
 ("Part station (40 %)",f"{L_MACHINE_YEAR*FUEL_P0*STATION_SHARE:,.0f} DT/an",f"≈ {L_MACHINE_YEAR*FUEL_P0*STATION_SHARE/12:,.0f} DT/mois"),
 ("(−) Leasing machine",f"-{MACHINE_LEASE_YEAR:,.0f} DT/an","57 928 @ 12 %/7 ans"),
 ("(−) Maintenance + internet",f"-{OPEX_MACHINE_MONTH*12:,.0f} DT/an",""),
 ("= Marge nette / machine",f"{L_MACHINE_YEAR*FUEL_P0*HCTECH_SHARE - MACHINE_LEASE_YEAR - OPEX_MACHINE_MONTH*12:,.0f} DT/an","avant frais de structure partagés"),
]
ws.append(["Poste","Valeur","Note"])
for i in range(1,4): style_header(ws.cell(row=3,column=i))
for row in unit:
    ws.append(list(row)); r=ws.max_row
    for cc in range(1,4): ws.cell(row=r,column=cc).border=border
    if row[0].startswith("="):
        for cc in range(1,4):
            ws.cell(row=r,column=cc).fill=PatternFill("solid",fgColor=LIGHT)
            ws.cell(row=r,column=cc).font=Font(bold=True,color=GREEN)
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=20; ws.column_dimensions["C"].width=34

# ---- Sheet 7: Actionnariat ----
ws=wb.create_sheet("Actionnariat")
title(ws,"Capital & Actionnariat","Capital social 300 000 DT — apporté intégralement par Nessim Mami")
ws.append([]); ws.append([])
ws.append(["Actionnaire","% parts","Apport cash","Valeur implicite des parts","Nature"])
for i in range(1,6): style_header(ws.cell(row=3,column=i))
post_money = CAPITAL/SHARES['Nessim Mami']
for name,pct in SHARES.items():
    cash=CASH_CONTRIB[name]
    val=pct*post_money
    nature="Apport cash" if cash>0 else "Apport industrie/nature (techno, réseau, gestion)"
    ws.append([name,pct,cash,val,nature]); r=ws.max_row
    ws.cell(row=r,column=2).number_format=PCT
    ws.cell(row=r,column=3).number_format=DTFMT2
    ws.cell(row=r,column=4).number_format=DTFMT2
    for cc in range(1,6): ws.cell(row=r,column=cc).border=border
ws.append(["TOTAL",1.0,sum(CASH_CONTRIB.values()),post_money,""]); r=ws.max_row
ws.cell(row=r,column=2).number_format=PCT; ws.cell(row=r,column=3).number_format=DTFMT2
ws.cell(row=r,column=4).number_format=DTFMT2
for cc in range(1,6):
    ws.cell(row=r,column=cc).font=Font(bold=True); ws.cell(row=r,column=cc).fill=PatternFill("solid",fgColor=LIGHT)
ws.append([])
ws.append(["Note : Nessim finance 100 % du capital (300 k) pour 50 %. Les parts d'Ali et Lamine"])
ws.append(["correspondent à des apports en industrie/nature (exclusivité GEKO, réseau pétrolier, gestion)"])
ws.append(["valorisés ~300 k DT pre-money. À formaliser dans le pacte d'actionnaires."])
ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=10
ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=22; ws.column_dimensions["E"].width=46

# ---- Sheet 8: Dividendes ----
ws=wb.create_sheet("Dividendes")
title(ws,"Dividendes — Option 1 (Nessim prioritaire)",
      "Nessim encaisse 100% jusqu'à capital + prime 20% (360k), puis pro-rata 50/25/25 · payout 70%")
ws.append([]); ws.append([])
ws.append(["Actionnaire"]+[f"An {y}" for y in range(1,8)]+["TOTAL 7 ans"])
for i in range(1,10): style_header(ws.cell(row=3,column=i))
def div_row(name, label, fill=None):
    vals=[DIV[y][name] for y in range(1,8)]; tot=sum(vals)
    ws.append([label]+vals+[tot]); r=ws.max_row
    for j in range(2,10):
        c=ws.cell(row=r,column=j); c.number_format=DTFMT2; c.border=border
    ws.cell(row=r,column=1).border=border
    if fill:
        for cc in range(1,10):
            ws.cell(row=r,column=cc).fill=PatternFill("solid",fgColor=fill)
            ws.cell(row=r,column=cc).font=Font(bold=True)
div_row("Nessim Mami","Nessim Mami (50 %)", fill=LIGHT)
div_row("Ali Ben Hamoud","Ali Ben Hamoud (25 %)")
div_row("Lamine","Lamine (25 %)")
ws.append(["TOTAL distribué"]+[DIV_TOTAL[y] for y in range(1,8)]+[sum(DIV_TOTAL.values())]); r=ws.max_row
for cc in range(1,10):
    ws.cell(row=r,column=cc).font=Font(bold=True); ws.cell(row=r,column=cc).number_format=DTFMT2
    ws.cell(row=r,column=cc).border=border
ws.append([])
ws.append([f"Nessim : capital remboursé fin An {NESSIM_PAYBACK} · TRI ≈ {NESSIM_IRR*100:.0f} % · "
           f"multiple {SH_TOTAL['Nessim Mami']/CAPITAL:.1f}× · total reçu {SH_TOTAL['Nessim Mami']:,.0f} DT"])
ws.cell(row=ws.max_row,column=1).font=Font(bold=True,color=GREEN)
ws.append(["Note : dividende déclaré au titre de l'exercice N, versé en N+1. 30 % du résultat conservé en réserve."])
ws.column_dimensions["A"].width=24
for col in range(2,10): ws.column_dimensions[get_column_letter(col)].width=13

# Cap table -> aussi le détail dividendes total par associé
ws_a=wb["Actionnariat"]
ws_a.append([])
ws_a.append(["Dividendes reçus sur 7 ans (Option 1) :"])
ws_a.cell(row=ws_a.max_row,column=1).font=Font(bold=True,color=DARK)
for name in SHARES:
    ws_a.append([name, "", "", SH_TOTAL[name], f"{SH_TOTAL[name]/CAPITAL:.1f}× le capital" if name=='Nessim Mami' else ""])
    ws_a.cell(row=ws_a.max_row,column=4).number_format=DTFMT2

wb.save("/Users/os/Desktop/HCTech/BP_HCTECH_7ans.xlsx")

# ════════════════════════ CONSOLE SUMMARY ════════════════════════
print("MACHINE_HT =", round(MACHINE_HT), "| Leasing/mois =", round(MACHINE_LEASE_MONTH,1),
      "| Leasing/an =", round(MACHINE_LEASE_YEAR))
print(f"{'':22}" + "".join(f"{'An'+str(y):>13}" for y in range(1,8)))
def pr(lbl, key):
    print(f"{lbl:22}" + "".join(f"{YA[y][key]:>13,.0f}" for y in range(1,8)))
print(f"{'Parc':22}" + "".join(f"{YA[y]['parc']:>13.0f}" for y in range(1,8)))
pr("CA HCTECH","rev"); pr("Leasing machines","mlease"); pr("Masse salariale","payroll")
pr("EBITDA/Rés.exploit.","ebitda"); pr("Impôt","tax"); pr("Résultat net","net")
print(f"{'Trésorerie EOY':22}" + "".join(f"{CASH_EOY[y]:>13,.0f}" for y in range(1,8)))
cum=0
print(f"{'Net cumulé':22}",end="")
for y in range(1,8):
    cum+=YA[y]['net']; print(f"{cum:>13,.0f}",end="")
print()
print("Trésorerie min (84 mois):", f"{MIN_CASH:,.0f} DT")
print("Marge nette/machine/an (an1):",
      f"{machine_rev_month(1)*12 - MACHINE_LEASE_YEAR - OPEX_MACHINE_MONTH*12:,.0f} DT")
print("Excel -> BP_HCTECH_7ans.xlsx")
