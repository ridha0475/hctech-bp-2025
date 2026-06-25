#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HCTECH — Dossier de financement LEASING (classeur Excel à FORMULES VIVANTES).
Toutes les données d'entrée sont sur l'onglet « Hypothèses » ; tout le reste est
calculé par des formules Excel qui se recalculent quand on change une hypothèse.
Public cible : sociétés de leasing (focus : capacité de remboursement / couverture).

Génère : HCTECH_BP_Leasing.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
LOGO = os.path.join(ROOT, "assets", "logo-hctech.png")

# ── Palette ───────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2E75B6"
LIGHT  = "D6E4F0"
LIGHT2 = "EAF1F9"
GREEN  = "1A7A4A"
GREENL = "D7EFE2"
AMBER  = "B8860B"
GREY   = "808080"
GREYL  = "F2F2F2"
WHITE  = "FFFFFF"
DARK   = "333333"

DTFMT  = '#,##0 "DT"'
DTFMT2 = '#,##0'
PCT    = '0.0%'
PCT0   = '0%'
MULT   = '0.0"×"'
NUM    = '#,##0'

thin  = Side(style="thin",  color="BBBBBB")
med   = Side(style="medium", color=BLUE)
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
TOPLINE    = Border(top=Side(style="medium", color=NAVY))

def fill(c): return PatternFill("solid", fgColor=c)

def style(cell, *, bold=False, size=10, color=DARK, bg=None, align=None,
          fmt=None, border=False, italic=False, wrap=False):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    if bg: cell.fill = fill(bg)
    if align: cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    else: cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    if fmt: cell.number_format = fmt
    if border: cell.border = BORDER

def sheet_title(ws, title, subtitle, last_col="J"):
    ws.merge_cells(f"B1:{last_col}1")
    style(ws["B1"], bold=True, size=16, color=NAVY)
    ws["B1"] = title
    ws.merge_cells(f"B2:{last_col}2")
    style(ws["B2"], size=10, color=GREY, italic=True)
    ws["B2"] = subtitle
    ws.row_dimensions[1].height = 22

def section(ws, row, text, last_col="J", color=BLUE):
    ws.merge_cells(f"B{row}:{last_col}{row}")
    style(ws[f"B{row}"], bold=True, size=11, color=WHITE, bg=color, align="left")
    ws[f"B{row}"] = "  " + text
    ws.row_dimensions[row].height = 18

def band(ws, cell_from, cell_to, text, color=BLUE):
    ws.merge_cells(f"{cell_from}:{cell_to}")
    style(ws[cell_from], bold=True, size=11, color=WHITE, bg=color, align="left")
    ws[cell_from] = "  " + text

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
#  ONGLET 2 — HYPOTHÈSES  (toutes les entrées — modifiables)
# ════════════════════════════════════════════════════════════════════
H = "'Hypothèses'"
def hyp(r): return f"{H}!$C${r}"

ws = wb.active
ws.title = "Hypothèses"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 46
sheet_title(ws, "Hypothèses du modèle",
            "Toutes les cellules bleues sont modifiables — le classeur entier se recalcule automatiquement.", "E")

# (row, label, value-or-formula, unit, note, is_input)
HROWS = [
  ("S", "PARAMÈTRES MACHINE & LEASING"),
  (5,  "Taux de change USD / TND",            2.95,   "",      "Cours de référence", True),
  (6,  "Prix machine (départ usine)",         18000,  "USD",   "Fournisseur GEKO (Corée)", True),
  (7,  "Frais d'import",                       0.025,  "%",     "Droits + transit", True),
  (8,  "Installation",                         3500,   "DT",    "Prestataire local", True),
  (9,  "Coût machine tout compris (HT)",      "=C6*C5*(1+C7)+C8", "DT", "Base de l'opération de leasing", False),
  (10, "Taux de leasing machine",              0.12,   "%/an",  "Marché tunisien (plafond usure 16,28 %)", True),
  (11, "Durée du leasing machine",             5,      "ans",   "Règle Tunisie : leasing équipement 5 ans", True),
  (12, "Mensualité leasing / machine",        "=C9*(C10/12)/(1-(1+C10/12)^(-C11*12))", "DT", "Annuité constante", False),
  (13, "Loyer leasing annuel / machine",      "=C12*12", "DT/an", "Pendant 5 ans, puis 0", False),
  (14, "Maintenance + internet / machine",     120,    "DT/mois","OPEX par machine", True),
  ("S", "REVENUS PAR MACHINE"),
  (17, "Volume station cible",                 12000,  "L/jour","Stations ≥ 12 000 L/j (~60 % du parc)", True),
  (18, "Taux de récupération",                 5,      "/1000", "Hypothèse médiane (fourchette 3–7/1000)", True),
  (19, "Essence récupérée / machine",         "=C17*C18/1000", "L/jour", "volume × taux", False),
  (20, "Prix de l'essence (An 1)",             2.525,  "DT/L",  "Prix à la pompe", True),
  (21, "Hausse du prix de l'essence",          0.045,  "%/an",  "CAGR réel 2016–2026", True),
  (22, "Part HCTECH (revenue share)",          0.60,   "%",     "Partage 60 / 40", True),
  (23, "Part station",                        "=1-C22","%",     "Station : 0 investissement", False),
  ("S", "DÉPLOIEMENT DU PARC"),
  (26, "Machines par vague",                   14,     "mach.", "Une vague tous les 6 mois", True),
  (27, "Intervalle entre vagues",              6,      "mois",  "", True),
  (28, "Mois du 1er déploiement",              3,      "mois",  "", True),
  (29, "Nombre de vagues",                     14,     "vagues","", True),
  (30, "Horizon du plan",                      84,     "mois",  "7 ans", True),
  (31, "Parc final",                          "=C26*C29", "mach.", "Cible à l'An 7", False),
  ("S", "CHARGES DE STRUCTURE"),
  (34, "Nombre de gérants / fondateurs",       2,      "pers.", "Ali + Lamine", True),
  (35, "Salaire fondateur (An 1)",             2000,   "DT/mois","", True),
  (36, "Croissance salaire fondateur",         0.50,   "%/an",  "Plafonnée", True),
  (37, "Plafond salaire fondateur",            6000,   "DT/mois","", True),
  (38, "Salaire technicien (An 1)",            1200,   "DT/mois","", True),
  (39, "Croissance salaire technicien",        0.07,   "%/an",  "", True),
  (40, "Machines par technicien",              48,     "mach.", "Capacité d'un technicien", True),
  (41, "Charges patronales",                   0.21,   "%",     "CNSS + TFP + FOPROLOS", True),
  (42, "Loyer bureau (An 1)",                  1200,   "DT/mois","", True),
  (43, "Croissance loyer bureau",              0.06,   "%/an",  "", True),
  (44, "Frais généraux (An 1)",                2000,   "DT/mois","Admin, déplacements, divers", True),
  (45, "Croissance frais généraux",            0.06,   "%/an",  "", True),
  ("S", "VÉHICULE UTILITAIRE (leasing)"),
  (48, "Prix du véhicule",                     70000,  "DT",    "Intervention / maintenance", True),
  (49, "Taux de leasing véhicule",             0.12,   "%/an",  "", True),
  (50, "Durée du leasing véhicule",            60,     "mois",  "5 ans", True),
  (51, "Mensualité véhicule",                 "=C48*(C49/12)/(1-(1+C49/12)^(-C50))", "DT", "", False),
  ("S", "FISCALITÉ & CAPITAL"),
  (54, "Impôt sur les sociétés (IS + CSS)",    0.28,   "%",     "IS 25 % + CSS 3 %", True),
  (55, "Capital social",                       300000, "DT",    "Apporté à 100 % par Nessim Mami", True),
  (56, "Payout dividendes",                    0.70,   "%",     "70 % distribué, 30 % en réserve", True),
  (57, "Prime de risque investisseur",         0.20,   "%",     "Sur récupération prioritaire", True),
  (58, "Récupération prioritaire Nessim",     "=C55*(1+C57)", "DT", "Capital + prime (= 360 000)", False),
  ("S", "RÉPARTITION DU CAPITAL"),
  (61, "Nessim Mami (investisseur)",           0.45,   "%",     "Apporte 100 % du capital", True),
  (62, "Ali Ben Hamoud (Directeur Technique)", 0.25,   "%",     "Apport industrie", True),
  (63, "Mohamed Lamine Belajouza (CEO)",       0.25,   "%",     "Apport industrie", True),
  (64, "Nazeh Ben Ammar (conseiller de Nessim)",0.05,  "%",     "Apport réseau institutionnel", True),
]
for item in HROWS:
    if item[0] == "S":
        # section header occupies the next available labelled row index implicitly via fixed rows
        continue
for item in HROWS:
    if item[0] == "S": continue
    r, label, val, unit, note, is_input = item
    style(ws[f"B{r}"], size=10, color=DARK)
    ws[f"B{r}"] = label
    cell = ws[f"C{r}"]
    cell.value = val
    is_pct = unit == "%"
    cell.number_format = (PCT if is_pct else (DTFMT2 if isinstance(val, (int, float)) and abs(val) >= 100 else '#,##0.000'))
    if isinstance(val, str) and val.startswith("="):
        style(cell, bold=True, color=NAVY, bg=GREYL, align="right", border=True,
              fmt=(PCT if is_pct else DTFMT2))
    else:
        style(cell, bold=True, color=NAVY, bg=LIGHT, align="right", border=True,
              fmt=cell.number_format)
    style(ws[f"D{r}"], size=9, color=GREY, align="left")
    ws[f"D{r}"] = unit
    style(ws[f"E{r}"], size=9, color=GREY, italic=True)
    ws[f"E{r}"] = note
# section bands
for r, txt in [(4,"PARAMÈTRES MACHINE & LEASING"),(16,"REVENUS PAR MACHINE"),(25,"DÉPLOIEMENT DU PARC"),
               (33,"CHARGES DE STRUCTURE"),(47,"VÉHICULE UTILITAIRE (leasing)"),(53,"FISCALITÉ & CAPITAL"),
               (60,"RÉPARTITION DU CAPITAL")]:
    section(ws, r, txt, "E")
ws.merge_cells("B66:E66")
style(ws["B66"], size=9, italic=True, color=GREY)
ws["B66"] = "Cellules bleues = saisie libre.  Cellules grisées = calculées (ne pas modifier)."

CAL = "'Calcul'"
def srng(col): return f"{CAL}!${col}$2:${col}$85"     # plage des 84 mois
def smif(col_val, y): return f"SUMIF({CAL}!$B$2:$B$85,{y},{srng(col_val)})"
def cal_year_end(col, y): return f"{CAL}!${col}${12*y+1}"   # valeur fin d'année (mois 12y -> ligne 12y+1)
def net_y(y): return f"({smif('R', y)}-MAX(0,{smif('R', y)})*{hyp(54)})"     # résultat net année y
def cumnet_body(y): return "+".join(net_y(k) for k in range(1, y+1))         # net cumulé jusqu'à y

# ════════════════════════════════════════════════════════════════════
#  ONGLET 4 — CALCUL (moteur mensuel, 84 mois) — formules
# ════════════════════════════════════════════════════════════════════
wc = wb.create_sheet("Calcul")
wc.sheet_view.showGridLines = False
heads = ["Mois","Année","Vagues cum.","Parc","Vagues>durée","Mach. en leasing","Prix essence",
         "Revenu HCTECH","OPEX mach.","Loyer leasing mach.","Nb tech.","Sal. fond.","Sal. tech.",
         "Masse sal.","Loyer bureau","Frais gén.","Leasing véh.","EBITDA mensuel","Impôt payé","Trésorerie cum."]
wc.merge_cells("A1:T1") if False else None
for j, h in enumerate(heads, start=1):
    c = wc.cell(row=1, column=j, value=h)
    style(c, bold=True, size=8, color=WHITE, bg=NAVY, align="center", wrap=True)
    wc.column_dimensions[get_column_letter(j)].width = 11
wc.row_dimensions[1].height = 30
for r in range(2, 86):
    m = r - 1
    wc[f"A{r}"] = m
    wc[f"B{r}"] = f"=INT((A{r}-1)/12)+1"
    wc[f"C{r}"] = (f"=IF(A{r}<{hyp(28)},0,MIN({hyp(29)},INT((A{r}-{hyp(28)})/{hyp(27)})+1))")
    wc[f"D{r}"] = f"=C{r}*{hyp(26)}"
    wc[f"E{r}"] = (f"=IF((A{r}-{hyp(11)}*12)<{hyp(28)},0,"
                  f"MIN({hyp(29)},INT((A{r}-{hyp(11)}*12-{hyp(28)})/{hyp(27)})+1))")
    wc[f"F{r}"] = f"=(C{r}-E{r})*{hyp(26)}"
    wc[f"G{r}"] = f"={hyp(20)}*(1+{hyp(21)})^(B{r}-1)"
    wc[f"H{r}"] = f"=D{r}*{hyp(19)}*(365/12)*G{r}*{hyp(22)}"
    wc[f"I{r}"] = f"=D{r}*{hyp(14)}"
    wc[f"J{r}"] = f"=F{r}*{hyp(12)}"
    wc[f"K{r}"] = f"=MAX(1,ROUNDUP(D{r}/{hyp(40)},0))"
    wc[f"L{r}"] = f"=MIN({hyp(35)}*(1+{hyp(36)})^(B{r}-1),{hyp(37)})"
    wc[f"M{r}"] = f"={hyp(38)}*(1+{hyp(39)})^(B{r}-1)"
    wc[f"N{r}"] = f"=({hyp(34)}*L{r}+K{r}*M{r})*(1+{hyp(41)})"
    wc[f"O{r}"] = f"={hyp(42)}*(1+{hyp(43)})^(B{r}-1)"
    wc[f"P{r}"] = f"={hyp(44)}*(1+{hyp(45)})^(B{r}-1)"
    wc[f"Q{r}"] = f"=IF(A{r}<={hyp(50)},{hyp(51)},0)"
    wc[f"R{r}"] = f"=H{r}-I{r}-J{r}-N{r}-O{r}-P{r}-Q{r}"
    wc[f"S{r}"] = (f"=IF(AND(MOD(A{r},12)=6,A{r}>=18),"
                  f"MAX(0,SUMIF($B$2:$B$85,B{r}-1,$R$2:$R$85))*{hyp(54)},0)")
    wc[f"T{r}"] = (f"={hyp(55)}+R{r}-S{r}" if r == 2 else f"=T{r-1}+R{r}-S{r}")
    for col in "ABCDEFGHIJKLMNOPQRST":
        cc = wc[f"{col}{r}"]
        cc.number_format = NUM if col in "ABCDEFK" else DTFMT2
        cc.font = Font(size=8, color=DARK)
wc.sheet_state = "hidden"   # moteur de calcul masqué (le livrable montre les synthèses)

# helper for year-labeled tables
YEARS = list(range(1, 8))
def year_header(ws, row, first_col=3, total=True):
    for i, y in enumerate(YEARS):
        c = ws.cell(row=row, column=first_col+i, value=f"An {y}")
        style(c, bold=True, size=10, color=WHITE, bg=BLUE, align="center")
    if total:
        c = ws.cell(row=row, column=first_col+7, value="Total 7 ans")
        style(c, bold=True, size=10, color=WHITE, bg=NAVY, align="center")

def line(ws, row, label, formula_for_year, *, fmt=DTFMT2, bold=False, bg=None,
         total="sum", indent=False, first_col=3, color=DARK):
    lab = ("   " if indent else "") + label
    style(ws.cell(row=row, column=2, value=lab), bold=bold, color=color, bg=bg, size=10)
    for i, y in enumerate(YEARS):
        col = first_col + i
        c = ws.cell(row=row, column=col, value=formula_for_year(y))
        style(c, bold=bold, color=color, bg=bg, fmt=fmt, border=True, align="right")
    if total is not None:
        cl = get_column_letter(first_col); cr = get_column_letter(first_col+6)
        tc = ws.cell(row=row, column=first_col+7)
        if total == "sum":
            tc.value = f"=SUM({cl}{row}:{cr}{row})"
        elif total == "last":
            tc.value = f"={get_column_letter(first_col+6)}{row}"
        else:
            tc.value = total(row)
        style(tc, bold=True, color=NAVY, bg=(bg or GREYL), fmt=fmt, border=True, align="right")

def setup_year_sheet(ws, title, subtitle):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    for i in range(3, 11):
        ws.column_dimensions[get_column_letter(i)].width = 13
    sheet_title(ws, title, subtitle, "J")

# ════════════════════════════════════════════════════════════════════
#  ONGLET — ÉTAT DE RÉSULTAT (présentation normée SCE Tunisie)
# ════════════════════════════════════════════════════════════════════
wr = wb.create_sheet("État de résultat")
setup_year_sheet(wr, "État de résultat prévisionnel",
                 "Présentation normée (Système Comptable des Entreprises — Tunisie). En DT — recalcul automatique.")
year_header(wr, 4)
style(wr.cell(row=5, column=2, value="Parc de machines en exploitation (fin d'année) — pour mémoire"),
      italic=True, size=9, color=GREY)
for i, y in enumerate(YEARS):
    cc = wr.cell(row=5, column=3+i, value=f"={cal_year_end('D', y)}")
    style(cc, italic=True, size=9, color=GREY, fmt=NUM, align="right")
band(wr, "B6", "J6", "PRODUITS D'EXPLOITATION", BLUE)
line(wr, 7, "Revenus (revenue share — part HCTECH 60 %)", lambda y: f"={smif('H', y)}", bold=True)
band(wr, "B8", "J8", "CHARGES D'EXPLOITATION", BLUE)
line(wr, 9, "Achats & charges externes — maintenance machines", lambda y: f"=-{smif('I', y)}", indent=True, color=GREY)
line(wr, 10, "Charges de leasing — machines", lambda y: f"=-{smif('J', y)}", indent=True, color=GREY)
line(wr, 11, "Charges de leasing — véhicule", lambda y: f"=-{smif('Q', y)}", indent=True, color=GREY)
line(wr, 12, "Charges de personnel", lambda y: f"=-{smif('N', y)}", indent=True, color=GREY)
line(wr, 13, "Loyer & charges locatives (bureau)", lambda y: f"=-{smif('O', y)}", indent=True, color=GREY)
line(wr, 14, "Autres charges d'exploitation", lambda y: f"=-{smif('P', y)}", indent=True, color=GREY)
line(wr, 15, "Total des charges d'exploitation",
     lambda y: f"=-({smif('I', y)}+{smif('J', y)}+{smif('Q', y)}+{smif('N', y)}+{smif('O', y)}+{smif('P', y)})", color=DARK)
line(wr, 16, "RÉSULTAT D'EXPLOITATION (= EBITDA, sans amortissement)", lambda y: f"={smif('R', y)}", bold=True, bg=LIGHT2)
line(wr, 17, "Charges financières nettes", lambda y: "=0", indent=True, color=GREY)
line(wr, 18, "Résultat des activités ordinaires avant impôt", lambda y: f"={smif('R', y)}", bold=True)
line(wr, 19, "Impôt sur les bénéfices (IS 25 % + CSS 3 %)", lambda y: f"=-MAX(0,{smif('R', y)})*{hyp(54)}", indent=True, color=GREY)
line(wr, 20, "RÉSULTAT NET DE L'EXERCICE", lambda y: f"={net_y(y)}", bold=True, bg=GREENL, color=GREEN)
line(wr, 21, "Résultat net cumulé", lambda y: "=" + cumnet_body(y), total="last", bold=True, color=NAVY)
wr.cell(row=23, column=2, value="Aucun amortissement de machine : le financement est porté par le leasing (location). "
        "Aucune charge financière bancaire : zéro emprunt.")
style(wr.cell(row=23, column=2), italic=True, size=9, color=GREY)
wr.merge_cells("B23:J23")

# ════════════════════════════════════════════════════════════════════
#  ONGLET — ÉTAT DE FLUX DE TRÉSORERIE (par activité, norme SCE)
# ════════════════════════════════════════════════════════════════════
wt = wb.create_sheet("Flux de trésorerie")
setup_year_sheet(wt, "État de flux de trésorerie prévisionnel",
                 "Présentation par activité (norme tunisienne). En DT — aucun flux d'investissement : les machines sont en leasing.")
year_header(wt, 4)
def colT(y): return get_column_letter(3+y-1)
band(wt, "B5", "J5", "FLUX LIÉS À L'EXPLOITATION", GREEN)
line(wt, 6, "EBITDA (capacité d'autofinancement)", lambda y: f"={smif('R', y)}", indent=True, color=GREY)
line(wt, 7, "(−) Impôt décaissé sur l'exercice", lambda y: f"=-{smif('S', y)}", indent=True, color=GREY)
line(wt, 8, "Flux net de trésorerie d'exploitation", lambda y: f"={colT(y)}6+{colT(y)}7", bold=True, bg=LIGHT2)
band(wt, "B9", "J9", "FLUX LIÉS À L'INVESTISSEMENT", GREEN)
line(wt, 10, "Acquisitions d'immobilisations (machines en leasing)", lambda y: "=0", indent=True, color=GREY)
line(wt, 11, "Flux net de trésorerie d'investissement", lambda y: "=0", bold=True, bg=LIGHT2)
band(wt, "B12", "J12", "FLUX LIÉS AU FINANCEMENT", GREEN)
line(wt, 13, "Apport en capital", lambda y: (f"={hyp(55)}" if y == 1 else "=0"), indent=True, color=GREY)
line(wt, 14, "Dividendes versés (au titre de l'exercice N-1)", lambda y: ("=0" if y == 1 else f"=-{hyp(56)}*{net_y(y-1)}"), indent=True, color=GREY)
line(wt, 15, "Flux net de trésorerie de financement", lambda y: f"={colT(y)}13+{colT(y)}14", bold=True, bg=LIGHT2)
line(wt, 16, "VARIATION DE TRÉSORERIE", lambda y: f"={colT(y)}8+{colT(y)}11+{colT(y)}15", bold=True)
line(wt, 17, "Trésorerie à l'ouverture", lambda y: ("=0" if y == 1 else f"={colT(y-1)}18"), indent=True, color=GREY, total=None)
line(wt, 18, "Trésorerie à la clôture", lambda y: f"={colT(y)}17+{colT(y)}16", bold=True, bg=GREENL, color=GREEN, total=None)
style(wt.cell(row=20, column=2, value="Trésorerie minimale sur 84 mois (avant distributions)"), bold=True, color=NAVY)
wt.cell(row=20, column=3, value=f"=MIN({CAL}!$T$2:$T$85)")
style(wt.cell(row=20, column=3), bold=True, color=GREEN, fmt=DTFMT, bg=GREENL, border=True, align="right")
wt.merge_cells("D20:J20")
style(wt.cell(row=20, column=4, value="→ le capital de 300 k suffit comme fonds de roulement, sans dette bancaire"),
      italic=True, size=9, color=GREY)

# ════════════════════════════════════════════════════════════════════
#  ONGLET — BILAN (présentation normée SCE Tunisie)
# ════════════════════════════════════════════════════════════════════
wb_ = wb.create_sheet("Bilan")
setup_year_sheet(wb_, "Bilan prévisionnel",
                 "Présentation normée (SCE Tunisie). En DT — machines en leasing = location simple (hors bilan). Équilibré par construction.")
year_header(wb_, 4, total=False)
def tresor_apresdistrib(y):
    base = f"{cal_year_end('T', y)}"
    if y == 1: return f"={base}"
    return f"={base}-{hyp(56)}*({cumnet_body(y-1)})"
def colY(y): return get_column_letter(3+y-1)
band(wb_, "B5", "J5", "ACTIFS", NAVY)
line(wb_, 6, "Actifs non courants (immobilisations)", lambda y: "=0", indent=True, color=GREY, total=None)
line(wb_, 7, "Actifs courants — trésorerie & équivalents", tresor_apresdistrib, indent=True, color=GREY, total=None)
line(wb_, 8, "TOTAL DES ACTIFS", lambda y: f"={colY(y)}6+{colY(y)}7", bold=True, bg=LIGHT2, total=None, color=NAVY)
band(wb_, "B9", "J9", "CAPITAUX PROPRES & PASSIFS", NAVY)
line(wb_, 10, "Capital social", lambda y: f"={hyp(55)}", indent=True, color=GREY, total=None)
line(wb_, 11, "Réserves & résultats reportés", lambda y: f"=(1-{hyp(56)})*({cumnet_body(y)})", indent=True, color=GREY, total=None)
line(wb_, 12, "Total des capitaux propres", lambda y: f"={colY(y)}10+{colY(y)}11", bold=True)
line(wb_, 13, "Passifs courants — dividendes à payer", lambda y: f"={hyp(56)}*{net_y(y)}", indent=True, color=GREY, total=None)
line(wb_, 14, "Passifs courants — impôt à payer (exigible N+1)", lambda y: f"=MAX(0,{smif('R', y)})*{hyp(54)}", indent=True, color=GREY, total=None)
line(wb_, 15, "Total des passifs courants", lambda y: f"={colY(y)}13+{colY(y)}14", color=DARK)
line(wb_, 16, "TOTAL CAPITAUX PROPRES & PASSIFS", lambda y: f"={colY(y)}12+{colY(y)}15", bold=True, bg=LIGHT2, total=None, color=NAVY)
line(wb_, 17, "Contrôle (Actif − Passif ≈ 0)", lambda y: f"={colY(y)}8-{colY(y)}16", fmt=DTFMT2, total=None, color=GREY)

# ════════════════════════════════════════════════════════════════════
#  ONGLET — NOTES AUX ÉTATS FINANCIERS
# ════════════════════════════════════════════════════════════════════
wn = wb.create_sheet("Notes")
wn.sheet_view.showGridLines = False
wn.column_dimensions["A"].width = 2
wn.column_dimensions["B"].width = 110
sheet_title(wn, "Notes aux états financiers prévisionnels",
            "Règles de présentation et hypothèses structurantes.", "B")
NOTES = [
  ("1. Référentiel comptable", "États financiers prévisionnels établis selon le Système Comptable des Entreprises (Tunisie), sur un horizon de 7 ans, exprimés en dinars tunisiens (DT)."),
  ("2. Traitement du leasing", "Les machines et le véhicule utilitaire sont financés en leasing. Traités en location simple : ils ne figurent pas à l'actif et les loyers sont comptabilisés en charges d'exploitation. Durée du leasing machine : 5 ans (au-delà, loyer nul)."),
  ("3. Absence de dette bancaire", "Le financement repose exclusivement sur le leasing et le capital social. Aucun emprunt bancaire — d'où des charges financières nulles."),
  ("4. Reconnaissance des revenus", "Modèle de revenue share 60/40 : HCTECH perçoit 60 % de la valeur de l'essence récupérée, la station 40 % (sans investissement). Revenus reconnus au rythme de la récupération."),
  ("5. Impôt sur les bénéfices", "IS 25 % + CSS 3 % = 28 % du résultat. L'impôt d'un exercice est exigible l'exercice suivant : il figure en passif courant au bilan et son décaissement apparaît l'année N+1 dans les flux."),
  ("6. Politique de distribution", "70 % du résultat distribué, 30 % conservé en réserve. L'investisseur (Nessim Mami) bénéficie d'une récupération prioritaire : capital + prime de risque de 20 % = 360 000 DT, avant tout partage au pro-rata."),
  ("7. Capital social", "300 000 DT, libéré à 100 % par l'investisseur Nessim Mami. Sert de fonds de roulement ; la trésorerie reste positive sur toute la période (minimum +278 k DT)."),
  ("8. Source des données", "Tous les montants découlent par formules de l'onglet « Hypothèses ». Modifier une hypothèse recalcule automatiquement l'ensemble des états financiers."),
]
rr = 4
for ttl, txt in NOTES:
    style(wn.cell(row=rr, column=2, value=ttl), bold=True, color=NAVY, size=11)
    rr += 1
    c = wn.cell(row=rr, column=2, value=txt)
    style(c, size=10, color=DARK, wrap=True)
    wn.row_dimensions[rr].height = 40
    rr += 2

# ════════════════════════════════════════════════════════════════════
#  ONGLET — RATIOS & INDICATEURS FINANCIERS
# ════════════════════════════════════════════════════════════════════
wf = wb.create_sheet("Ratios financiers")
setup_year_sheet(wf, "Ratios & indicateurs financiers",
                 "Agrégats et ratios usuels d'analyse crédit (EBITDA, marges, couverture, rentabilité). Recalcul automatique.")
year_header(wf, 4)
def cp(y): return f"({hyp(55)}+(1-{hyp(56)})*({cumnet_body(y)}))"          # capitaux propres
def tresor_clot(y):
    base = cal_year_end('T', y)
    return base if y == 1 else f"{base}-{hyp(56)}*({cumnet_body(y-1)})"
band(wf, "B5", "J5", "SOLDES & AGRÉGATS", BLUE)
line(wf, 6, "Chiffre d'affaires (revenus)", lambda y: f"={smif('H', y)}", bold=True)
line(wf, 7, "Croissance du chiffre d'affaires", lambda y: ("=\"—\"" if y == 1 else f"={smif('H', y)}/{smif('H', y-1)}-1"), fmt=PCT, indent=True, color=GREY, total=None)
line(wf, 8, "EBITDA (excédent brut d'exploitation)", lambda y: f"={smif('R', y)}", bold=True, bg=LIGHT2)
line(wf, 9, "EBIT (résultat d'exploitation)", lambda y: f"={smif('R', y)}", indent=True, color=GREY)
line(wf, 10, "Capacité d'autofinancement (CAF)", lambda y: f"={net_y(y)}", indent=True, color=GREY)
line(wf, 11, "Résultat net", lambda y: f"={net_y(y)}", bold=True, bg=GREENL, color=GREEN)
band(wf, "B12", "J12", "MARGES", BLUE)
line(wf, 13, "Marge d'EBITDA (% du CA)", lambda y: f"={smif('R', y)}/{smif('H', y)}", fmt=PCT, total=None)
line(wf, 14, "Marge nette (% du CA)", lambda y: f"={net_y(y)}/{smif('H', y)}", fmt=PCT, total=None)
band(wf, "B15", "J15", "COUVERTURE DU SERVICE LEASING (clé bailleur)", GREEN)
line(wf, 16, "Engagements leasing (loyers machines + véhicule)", lambda y: f"={smif('J', y)}+{smif('Q', y)}", bold=True)
line(wf, 17, "DSCR (EBITDA avant loyers / loyers)",
     lambda y: f"=({smif('R', y)}+{smif('J', y)}+{smif('Q', y)})/({smif('J', y)}+{smif('Q', y)})",
     fmt=MULT, bold=True, bg=GREENL, color=GREEN, total=lambda row: f"=MIN(C{row}:I{row})")
line(wf, 18, "EBITDA / loyers de leasing", lambda y: f"={smif('R', y)}/({smif('J', y)}+{smif('Q', y)})", fmt=MULT, indent=True, color=GREY, total=None)
band(wf, "B19", "J19", "STRUCTURE & RENTABILITÉ", BLUE)
line(wf, 20, "Capitaux propres", lambda y: f"={cp(y)}", total=None)
line(wf, 21, "Dette financière bancaire", lambda y: "=0", indent=True, color=GREY, total=None)
line(wf, 22, "Gearing (dette / capitaux propres)", lambda y: "=0", fmt=PCT, indent=True, color=GREY, total=None)
line(wf, 23, "Trésorerie nette de clôture", lambda y: f"={tresor_clot(y)}", total=None)
line(wf, 24, "ROE (résultat net / capitaux propres)", lambda y: f"={net_y(y)}/{cp(y)}", fmt=PCT, bold=True, total=None)
line(wf, 25, "Rentabilité du capital social", lambda y: f"={net_y(y)}/{hyp(55)}", fmt=PCT, indent=True, color=GREY, total=None)

# ════════════════════════════════════════════════════════════════════
#  ONGLET 8 — CAPACITÉ DE REMBOURSEMENT LEASING  (clé pour le bailleur)
# ════════════════════════════════════════════════════════════════════
wl = wb.create_sheet("Capacité leasing")
setup_year_sheet(wl, "Capacité de remboursement du leasing",
                 "L'indicateur clé pour le bailleur : les loyers sont couverts plusieurs fois par les flux d'exploitation.")
# unit economics band
section(wl, 4, "AU NIVEAU D'UNE MACHINE (An 1)", "J", GREEN)
unit = [
  (5, "Revenu HCTECH / machine / an (part 60 %)", f"={hyp(19)}*365*{hyp(20)}*{hyp(22)}", DTFMT),
  (6, "Loyer de leasing / machine / an", f"={hyp(13)}", DTFMT),
  (7, "Couverture du loyer par le revenu", f"=C5/C6", MULT),
]
for r, lab, frm, fmt in unit:
    style(wl.cell(row=r, column=2, value=lab), size=10, bold=(r == 7))
    c = wl.cell(row=r, column=3, value=frm)
    style(c, bold=True, color=(GREEN if r == 7 else NAVY), bg=(GREENL if r == 7 else LIGHT2),
          fmt=fmt, border=True, align="right")
    wl.merge_cells(f"D{r}:F{r}")
style(wl.cell(row=7, column=4, value="→ chaque machine couvre largement son propre loyer"), italic=True, size=9, color=GREY)

section(wl, 9, "AU NIVEAU DE LA SOCIÉTÉ (par année)", "J", GREEN)
year_header(wl, 10)
line(wl, 11, "Loyers leasing machines", lambda y: f"={smif('J', y)}", total="sum")
line(wl, 12, "Loyer leasing véhicule", lambda y: f"={smif('Q', y)}", total="sum")
line(wl, 13, "Total engagements leasing", lambda y: f"={smif('J', y)}+{smif('Q', y)}", bold=True, bg=LIGHT2)
line(wl, 14, "EBITDA avant loyers de leasing", lambda y: f"={smif('R', y)}+{smif('J', y)}+{smif('Q', y)}", bold=True)
line(wl, 15, "Couverture du service leasing (DSCR)",
     lambda y: f"=({smif('R', y)}+{smif('J', y)}+{smif('Q', y)})/({smif('J', y)}+{smif('Q', y)})",
     fmt=MULT, bold=True, bg=GREENL, color=GREEN, total=lambda row: f"=AVERAGE(C{row}:I{row})")
style(wl.cell(row=17, column=2), bold=True, color=NAVY)
wl.cell(row=17, column=2, value="DSCR minimum sur la période")
wl.cell(row=17, column=3, value="=MIN(C15:I15)")
style(wl.cell(row=17, column=3), bold=True, color=GREEN, bg=GREENL, fmt=MULT, border=True, align="right")
wl.merge_cells("D17:J17")
style(wl.cell(row=17, column=4), italic=True, size=9, color=GREY)
wl.cell(row=17, column=4, value="Un DSCR > 1 signifie que l'exploitation couvre intégralement le service du leasing chaque année.")

# ════════════════════════════════════════════════════════════════════
#  ONGLET 9 — DIVIDENDES & ACTIONNARIAT
# ════════════════════════════════════════════════════════════════════
wd = wb.create_sheet("Dividendes")
setup_year_sheet(wd, "Distribution & actionnariat",
                 "Priorité à l'investisseur (capital + prime 20 % = 360 k), puis pro-rata. 30 % conservé en réserve.")
year_header(wd, 4)
line(wd, 5, "Résultat net", lambda y: f"={net_y(y)}", bold=True)
line(wd, 6, "Enveloppe distribuable (payout 70 %)", lambda y: f"={hyp(56)}*{net_y(y)}", indent=True, color=GREY)
# priorité cumulée
def cumpool_before(y):  # somme des pools des années précédentes
    if y == 1: return "0"
    return "+".join(f"{hyp(56)}*{net_y(k)}" for k in range(1, y))
line(wd, 7, "Récupération prioritaire Nessim", lambda y: f"=MIN({hyp(56)}*{net_y(y)},MAX(0,{hyp(58)}-({cumpool_before(y)})))",
     indent=True, color=GREY)
line(wd, 8, "Reste à répartir au pro-rata", lambda y: f"={hyp(56)}*{net_y(y)}-{get_column_letter(3+y-1)}7", indent=True, color=GREY)
section(wd, 9, "DIVIDENDES PAR ASSOCIÉ (mêmes colonnes An 1 … An 7)", "J", BLUE)
col = lambda y: get_column_letter(3+y-1)
line(wd, 11, "Nessim Mami (45 %)", lambda y: f"={col(y)}7+{col(y)}8*{hyp(61)}", bold=True)
line(wd, 12, "Ali Ben Hamoud (25 %)", lambda y: f"={col(y)}8*{hyp(62)}")
line(wd, 13, "Mohamed Lamine Belajouza (25 %)", lambda y: f"={col(y)}8*{hyp(63)}")
line(wd, 14, "Nazeh Ben Ammar (5 %)", lambda y: f"={col(y)}8*{hyp(64)}")
line(wd, 15, "TOTAL distribué", lambda y: f"=SUM({col(y)}11:{col(y)}14)", bold=True, bg=LIGHT2, color=NAVY)
# multiple & TRI Nessim
style(wd.cell(row=17, column=2, value="Multiple Nessim (reçu / capital)"), bold=True, color=NAVY)
wd.cell(row=17, column=3, value=f"=J11/{hyp(55)}")
style(wd.cell(row=17, column=3), bold=True, color=GREEN, fmt=MULT, bg=GREENL, border=True, align="right")
style(wd.cell(row=18, column=2, value="TRI de l'investisseur (Nessim)"), bold=True, color=NAVY)
wd.cell(row=18, column=3, value="=IRR(C20:J20)")
style(wd.cell(row=18, column=3), bold=True, color=GREEN, fmt=PCT0, bg=GREENL, border=True, align="right")
# flux pour IRR : -capital en An0 (col C), puis Nessim An1..7
style(wd.cell(row=20, column=2, value="Flux investisseur (pour TRI)"), size=9, color=GREY, italic=True)
wd.cell(row=20, column=3, value=f"=-{hyp(55)}")
style(wd.cell(row=20, column=3), size=9, color=GREY, fmt=DTFMT2)
for y in YEARS:
    c = wd.cell(row=20, column=3+y, value=f"={get_column_letter(3+y-1)}11")
    style(c, size=9, color=GREY, fmt=DTFMT2)

# ════════════════════════════════════════════════════════════════════
#  ONGLET 3 — ÉCONOMIE D'UNE MACHINE
# ════════════════════════════════════════════════════════════════════
we = wb.create_sheet("Économie machine")
we.sheet_view.showGridLines = False
we.column_dimensions["A"].width = 2
we.column_dimensions["B"].width = 44
we.column_dimensions["C"].width = 16
we.column_dimensions["D"].width = 4
we.column_dimensions["E"].width = 46
sheet_title(we, "Économie d'une machine", "Le cœur du modèle : chaque machine s'autofinance et dégage une marge nette.", "E")
emrows = [
  (4, "INVESTISSEMENT & FINANCEMENT", None, None, True),
  (5, "Coût machine tout compris (HT)", f"={hyp(9)}", DTFMT, False),
  (6, "Financement", None, None, False, "Leasing 100 % (0 cash mobilisé par HCTECH)"),
  (7, "Loyer de leasing / mois", f"={hyp(12)}", DTFMT, False),
  (8, "Loyer de leasing / an (pendant 5 ans)", f"={hyp(13)}", DTFMT, False),
  (10, "REVENUS (scénario réaliste 5/1000)", None, None, True),
  (11, "Essence récupérée / an", f"={hyp(19)}*365", '#,##0 "L"', False),
  (12, "Valeur brute récupérée / an", f"={hyp(19)}*365*{hyp(20)}", DTFMT, False),
  (13, "Part HCTECH (60 %)", f"={hyp(19)}*365*{hyp(20)}*{hyp(22)}", DTFMT, False),
  (14, "Part station (40 %)", f"={hyp(19)}*365*{hyp(20)}*{hyp(23)}", DTFMT, False),
  (16, "MARGE PAR MACHINE", None, None, True),
  (17, "(−) Loyer leasing / an", f"=-{hyp(13)}", DTFMT, False),
  (18, "(−) Maintenance + internet / an", f"=-{hyp(14)}*12", DTFMT, False),
  (19, "Marge nette PENDANT le leasing (5 ans)", f"={hyp(19)}*365*{hyp(20)}*{hyp(22)}-{hyp(13)}-{hyp(14)}*12", DTFMT, "green"),
  (20, "Marge nette APRÈS payoff (An 6-7, loyer = 0)", f"={hyp(19)}*365*{hyp(20)}*{hyp(22)}-{hyp(14)}*12", DTFMT, "green"),
  (21, "Couverture du loyer par la part HCTECH", f"={hyp(19)}*365*{hyp(20)}*{hyp(22)}/{hyp(13)}", MULT, "blue"),
]
for item in emrows:
    r, lab = item[0], item[1]
    if item[4] is True:
        section(we, r, lab, "E", BLUE); continue
    style(we.cell(row=r, column=2, value=lab), size=10, bold=(item[4] in ("green","blue")))
    if item[2]:
        c = we.cell(row=r, column=3, value=item[2])
        kind = item[4]
        style(c, bold=True,
              color=(GREEN if kind == "green" else (BLUE if kind == "blue" else NAVY)),
              bg=(GREENL if kind == "green" else (LIGHT if kind == "blue" else LIGHT2)),
              fmt=item[3], border=True, align="right")
    if len(item) > 5:
        style(we.cell(row=r, column=5, value=item[5]), italic=True, size=9, color=GREY)

# ════════════════════════════════════════════════════════════════════
#  ONGLET 10 — SYNTHÈSE (tableau de bord)
# ════════════════════════════════════════════════════════════════════
wy = wb.create_sheet("Synthèse")
wy.sheet_view.showGridLines = False
for col, w in [("A",2),("B",40),("C",18),("D",4),("E",40),("F",18)]:
    wy.column_dimensions[col].width = w
sheet_title(wy, "Synthèse — indicateurs clés", "Vue d'ensemble du dossier de financement leasing.", "F")
band(wy, "B4", "C4", "EXPLOITATION", BLUE)
# left column KPIs
kpis_left = [
  (5, "Parc de machines (An 7)", f"={cal_year_end('D',7)}", NUM),
  (6, "Chiffre d'affaires HCTECH (An 7)", f"='État de résultat'!I7", DTFMT),
  (7, "EBITDA (An 7)", "='Ratios financiers'!I8", DTFMT),
  (8, "Marge d'EBITDA (An 7)", "='Ratios financiers'!I13", PCT),
  (9, "Résultat net (An 7)", f"='État de résultat'!I20", DTFMT),
  (10, "Résultat net cumulé (7 ans)", f"='État de résultat'!J20", DTFMT),
  (11, "Trésorerie minimale (84 mois)", f"='Flux de trésorerie'!C20", DTFMT),
]
for r, lab, frm, fmt in kpis_left:
    style(wy.cell(row=r, column=2, value=lab), size=10)
    c = wy.cell(row=r, column=3, value=frm)
    style(c, bold=True, color=NAVY, bg=LIGHT2, fmt=fmt, border=True, align="right")
band(wy, "B12", "C12", "ATTRACTIVITÉ POUR LE BAILLEUR", GREEN)
kpis_l2 = [
  (13, "Couverture du loyer / machine", "='Capacité leasing'!C7", MULT),
  (14, "DSCR minimum (société)", "='Capacité leasing'!C17", MULT),
  (15, "Total engagements leasing (7 ans)", "='Capacité leasing'!J13", DTFMT),
  (16, "Dette bancaire", "=0", DTFMT),
]
for r, lab, frm, fmt in kpis_l2:
    style(wy.cell(row=r, column=2, value=lab), size=10)
    c = wy.cell(row=r, column=3, value=frm)
    style(c, bold=True, color=GREEN, bg=GREENL, fmt=fmt, border=True, align="right")
# right column : actionnariat & retours
band(wy, "E4", "F4", "ACTIONNARIAT & RETOURS", NAVY)
wy.cell(row=5, column=5, value="Capital social"); wy.cell(row=5, column=6, value=f"={hyp(55)}")
wy.cell(row=6, column=5, value="Total distribué (7 ans)"); wy.cell(row=6, column=6, value="=Dividendes!J15")
wy.cell(row=7, column=5, value="Dividendes Nessim (7 ans)"); wy.cell(row=7, column=6, value="=Dividendes!J11")
wy.cell(row=8, column=5, value="Multiple Nessim"); wy.cell(row=8, column=6, value="=Dividendes!C17")
wy.cell(row=9, column=5, value="TRI investisseur (Nessim)"); wy.cell(row=9, column=6, value="=Dividendes!C18")
for r in range(5, 10):
    style(wy.cell(row=r, column=5), size=10)
    fmt = MULT if r == 8 else (PCT0 if r == 9 else DTFMT)
    style(wy.cell(row=r, column=6), bold=True, color=NAVY, bg=LIGHT2, fmt=fmt, border=True, align="right")
band(wy, "E12", "F12", "RÉPARTITION DU CAPITAL", NAVY)
captable = [(13,"Nessim Mami (investisseur)",61),(14,"Ali Ben Hamoud (DT)",62),
            (15,"Mohamed Lamine Belajouza (CEO)",63),(16,"Nazeh Ben Ammar (conseiller)",64)]
for r, lab, hr in captable:
    style(wy.cell(row=r, column=5, value=lab), size=10)
    c = wy.cell(row=r, column=6, value=f"={hyp(hr)}")
    style(c, bold=True, color=NAVY, bg=LIGHT2, fmt=PCT0, border=True, align="right")

# ════════════════════════════════════════════════════════════════════
#  ONGLET 1 — PAGE DE GARDE
# ════════════════════════════════════════════════════════════════════
wg = wb.create_sheet("Garde")
wg.sheet_view.showGridLines = False
for col, w in [("A",3),("B",30),("C",30),("D",30),("E",6)]:
    wg.column_dimensions[col].width = w
if os.path.exists(LOGO):
    img = XLImage(LOGO); img.width = 200; img.height = 93
    wg.add_image(img, "B2")
wg.merge_cells("B8:D8"); style(wg["B8"], bold=True, size=26, color=NAVY, align="center")
wg["B8"] = "HCTech SARL"
wg.merge_cells("B9:D9"); style(wg["B9"], size=13, color=BLUE, align="center")
wg["B9"] = "Récupération des vapeurs d'essence en station-service"
wg.merge_cells("B11:D11"); style(wg["B11"], bold=True, size=16, color=GREEN, align="center")
wg["B11"] = "Business Plan & Dossier de financement — Leasing"
wg.merge_cells("B12:D12"); style(wg["B12"], size=11, color=GREY, align="center", italic=True)
wg["B12"] = "Modèle Revenue Share 60 / 40 · Machines financées en leasing · Horizon 7 ans"
# highlights box
hb = [
  (15, "Parc cible (An 7)", "='Synthèse'!C5", NUM),
  (16, "Résultat net cumulé (7 ans)", "='Synthèse'!C8", DTFMT),
  (17, "Couverture du loyer / machine", "='Synthèse'!C13", MULT),
  (18, "DSCR minimum", "='Synthèse'!C14", MULT),
  (19, "Trésorerie minimale", "='Synthèse'!C9", DTFMT),
  (20, "Dette bancaire", "=0", DTFMT),
]
style(wg["B14"], bold=True, size=12, color=NAVY); wg["B14"] = "Points clés"
for r, lab, frm, fmt in hb:
    style(wg.cell(row=r, column=2, value=lab), size=11, color=DARK)
    wg.merge_cells(f"C{r}:D{r}")
    c = wg.cell(row=r, column=3, value=frm)
    style(c, bold=True, size=12, color=NAVY, bg=LIGHT2, fmt=fmt, border=True, align="right")
wg.merge_cells("B23:D23"); style(wg["B23"], size=10, color=DARK)
wg["B23"] = "Société à Responsabilité Limitée — RNE : 1953633L — 01 av. Hedi Khefacha, 2080 Ariana"
wg.merge_cells("B25:D25"); style(wg["B25"], size=9, italic=True, color=GREY, align="center")
wg["B25"] = "Document confidentiel — réservé aux établissements de leasing et partenaires financiers identifiés."
wg.merge_cells("B26:D26"); style(wg["B26"], size=9, italic=True, color=GREY, align="center")
wg["B26"] = "Préparé par Mohamed Ridha Belajouza, consultant et conseiller stratégique."

# ── ordre des onglets ────────────────────────────────────
order = ["Garde","Hypothèses","Économie machine","État de résultat","Bilan","Flux de trésorerie",
         "Ratios financiers","Capacité leasing","Dividendes","Notes","Synthèse","Calcul"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = wb["Garde"]

OUT = os.path.join(ROOT, "HCTECH_BP_Leasing.xlsx")
wb.save(OUT)
print("OK ->", OUT)
